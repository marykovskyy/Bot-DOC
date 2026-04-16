from __future__ import annotations
import time
import os
import asyncio
import tempfile
import random
import logging
import pandas as pd
import threading
from typing import Optional, List, Set, Dict, Any

from config import SCRAPER_CONFIG
from constants import (
    CAPTCHA_MAX_WAIT_SEC, CAPTCHA_POLL_INTERVAL_SEC,
    ELEMENT_WAIT_RETRIES, BROWSER_LAUNCH_TIMEOUT_SEC,
    SHEETS_WRITE_DELAY,
)
from DrissionPage import ChromiumPage, ChromiumOptions  # type: ignore[import]
import proxy.manager as proxy_manager
from scrapers.california import scrape_california
from scrapers.denmark import scrape_denmark
from scrapers.czech import scrape_czech
from scrapers.uk_api import scrape_uk_api
from scrapers.latvia import scrape_latvia
from scrapers.new_zealand import scrape_new_zealand
from scrapers.thailand import scrape_thailand
from scrapers.france import scrape_france_api   # ← API-скрапер (pappers.ai JSON)
from scrapers.finland import scrape_finland_api  # ← API-скрапер (PRH open data)
from scrapers import turkey as turkey_scraper
import database
import gsheets

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
#  ЧЕРГА ЗАПИСУ В GOOGLE SHEETS
#  Один фоновий потік пише послідовно з паузою —
#  уникає 429 і гарантує що всі записи дійдуть
# ─────────────────────────────────────────────
import queue

_sheets_queue: queue.Queue = queue.Queue()
_SHEETS_WRITE_DELAY = SHEETS_WRITE_DELAY  # секунд між записами (ліміт Sheets ~60 req/min)


def _sheets_worker() -> None:
    """Фоновий потік: бере завдання з черги і пише в Sheets по одному."""
    while True:
        task = _sheets_queue.get()
        if task is None:  # сигнал зупинки
            _sheets_queue.task_done()
            break
        name, link, site_key = task
        try:
            gsheets.append_to_sheet(name, link, site_key)
        except Exception as e:
            logger.error("Sheets worker помилка: %s", e)
        finally:
            _sheets_queue.task_done()
            time.sleep(_SHEETS_WRITE_DELAY)


# Запускаємо один постійний потік-воркер (daemon — не блокує завершення процесу)
_sheets_thread = threading.Thread(target=_sheets_worker, daemon=True)
_sheets_thread.start()


def _enqueue_sheet_write(name: str, link: str, site_key: str) -> None:
    """Ставить запис у чергу. Повертається одразу, не блокує скрапер."""
    _sheets_queue.put((name, link, site_key))


def flush_sheets_queue() -> None:
    """Чекає поки всі записи в чергу будуть оброблені."""
    _sheets_queue.join()


# --- Таймаути (константи замість магічних чисел) ---
# Таймаути — визначені в constants.py, тут лише підтвердження що імпортовані
# CAPTCHA_MAX_WAIT_SEC, CAPTCHA_POLL_INTERVAL_SEC, ELEMENT_WAIT_RETRIES, BROWSER_LAUNCH_TIMEOUT_SEC



# ─────────────────────────────────────────────
#  LOCAL PROXY RELAY
#  Chrome → localhost:PORT (без пароля)
#  → реальний проксі (з Basic Auth)
#
#  Чому цей підхід:
#  - Chrome 130+ вимкнув Manifest V2 extensions
#  - Chrome не підтримує user:pass@ в --proxy-server
#  - Relay вирішує обидві проблеми: Chrome бачить
#    localhost без авторизації, relay сам додає Auth
# ─────────────────────────────────────────────

class LocalProxyRelay:
    """
    Asyncio HTTP/HTTPS proxy relay з Basic Auth до upstream.

    Chrome → 127.0.0.1:LOCAL_PORT (без пароля)
           → upstream:port (з Proxy-Authorization header)

    Ключові виправлення v3:
    - _pipe НЕ закриває writer — тільки _handle закриває у finally
    - asyncio.wait(FIRST_COMPLETED) замість gather — коли один напрямок
      закривається, другий скасовується через cancel(), а не обривається
    - Правильне читання CONNECT-відповіді від upstream (loop до \r\n\r\n)
    - _start_event синхронізує старт сервера перед поверненням start()
    """

    def __init__(self, upstream_host: str, upstream_port: int,
                 username: str, password: str):
        self.upstream_host = upstream_host
        self.upstream_port = upstream_port
        import base64 as _b64
        creds = _b64.b64encode(f"{username}:{password}".encode()).decode()
        self._auth_header = (
            b"Proxy-Authorization: Basic " + creds.encode() + b"\r\n"
        )
        self.local_port = self._free_port()
        self._loop: Optional[asyncio.AbstractEventLoop] = None
        self._server = None

    @staticmethod
    def _free_port() -> int:
        import socket as _s
        with _s.socket() as s:
            s.bind(("127.0.0.1", 0))
            return s.getsockname()[1]

    def _inject_auth(self, data: bytes) -> bytes:
        """Вставляє Proxy-Authorization після першого \r\n\r\n якщо ще немає."""
        if b"Proxy-Authorization" not in data:
            return data.replace(
                b"\r\n\r\n",
                b"\r\n" + self._auth_header + b"\r\n",
                1
            )
        return data

    @staticmethod
    async def _pipe(reader: asyncio.StreamReader,
                    writer: asyncio.StreamWriter) -> None:
        """
        Одностороннє пересилання reader → writer.
        НЕ закриває writer — це робить _handle у finally.
        """
        try:
            while True:
                data = await reader.read(65536)
                if not data:
                    return
                writer.write(data)
                await writer.drain()
        except (ConnectionResetError, BrokenPipeError,
                asyncio.CancelledError, OSError):
            pass
        except Exception:
            pass

    @staticmethod
    async def _read_until_blank_line(reader: asyncio.StreamReader,
                                     timeout: float = 15.0) -> bytes:
        """Читає HTTP заголовки до \r\n\r\n з таймаутом."""
        buf = b""
        while b"\r\n\r\n" not in buf:
            chunk = await asyncio.wait_for(reader.read(4096), timeout=timeout)
            if not chunk:
                break
            buf += chunk
        return buf

    async def _relay_bidirectional(
        self,
        r1: asyncio.StreamReader, w1: asyncio.StreamWriter,
        r2: asyncio.StreamReader, w2: asyncio.StreamWriter,
    ) -> None:
        """
        Двосторонній relay: r1→w2 і r2→w1 одночасно.
        Коли один напрямок закривається — другий скасовується.
        Самі writer'и НЕ закриваються тут.
        """
        t1 = asyncio.create_task(self._pipe(r1, w2))
        t2 = asyncio.create_task(self._pipe(r2, w1))
        try:
            done, pending = await asyncio.wait(
                [t1, t2], return_when=asyncio.FIRST_COMPLETED
            )
            for task in pending:
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    pass
        except Exception:
            t1.cancel()
            t2.cancel()

    async def _handle(self,
                      client_r: asyncio.StreamReader,
                      client_w: asyncio.StreamWriter) -> None:
        upstream_r: Optional[asyncio.StreamReader] = None
        upstream_w: Optional[asyncio.StreamWriter] = None
        try:
            # 1. Читаємо заголовки від Chrome
            head = await self._read_until_blank_line(client_r, timeout=30.0)
            if not head:
                return

            method = head.split(b" ")[0].upper()

            # 2. Підключаємось до upstream проксі
            upstream_r, upstream_w = await asyncio.wait_for(
                asyncio.open_connection(self.upstream_host, self.upstream_port),
                timeout=15.0
            )

            # 3. Пересилаємо запит з авторизацією
            upstream_w.write(self._inject_auth(head))
            await upstream_w.drain()

            if method == b"CONNECT":
                # HTTPS тунель: читаємо відповідь "200 Connection established"
                resp = await self._read_until_blank_line(upstream_r, timeout=15.0)

                # Пересилаємо відповідь Chrome
                client_w.write(resp)
                await client_w.drain()

                # Якщо upstream відмовив — виходимо
                status_line = resp.split(b"\r\n")[0]
                if b"200" not in status_line:
                    logger.warning("Upstream proxy відмовив CONNECT: %s",
                                   status_line.decode(errors="replace"))
                    return

            # 4. Двосторонній relay (HTTP і HTTPS)
            await self._relay_bidirectional(client_r, client_w,
                                            upstream_r, upstream_w)

        except asyncio.TimeoutError:
            logger.debug("Proxy relay timeout")
        except Exception as exc:
            logger.debug("Proxy relay error: %s", exc)
        finally:
            for w in (upstream_w, client_w):
                if w is not None:
                    try:
                        w.close()
                        await asyncio.wait_for(w.wait_closed(), timeout=2.0)
                    except Exception:
                        pass

    def _run_loop(self, started_event: threading.Event) -> None:
        """Asyncio event loop у фоновому потоці."""
        self._loop = asyncio.new_event_loop()
        asyncio.set_event_loop(self._loop)

        async def _boot():
            self._server = await asyncio.start_server(
                self._handle, "127.0.0.1", self.local_port
            )
            started_event.set()   # сигналізуємо що сервер готовий
            await self._server.serve_forever()

        self._loop.run_until_complete(_boot())

    def start(self) -> int:
        """Запускає relay і чекає поки він буде готовий. Повертає порт."""
        ready = threading.Event()
        t = threading.Thread(target=self._run_loop, args=(ready,), daemon=True)
        t.start()
        # Чекаємо реального старту сервера (не просто sleep)
        if not ready.wait(timeout=5.0):
            raise RuntimeError("LocalProxyRelay не запустився за 5 секунд")
        logger.info("LocalProxyRelay запущено: 127.0.0.1:%d → %s:%d",
                    self.local_port, self.upstream_host, self.upstream_port)
        return self.local_port

    def stop(self) -> None:
        """Зупиняє relay."""
        if self._loop and self._loop.is_running():
            if self._server:
                self._loop.call_soon_threadsafe(self._server.close)
            self._loop.call_soon_threadsafe(self._loop.stop)


def get_page(chat_id: int, status_dict: dict, site_key: str = "General") -> tuple[Optional[ChromiumPage], Optional["LocalProxyRelay"]]:
    """
    Запускає Chrome з проксі через LocalProxyRelay.

    Схема: Chrome → localhost:PORT (без auth) → LocalProxyRelay → реальний проксі (з Basic Auth)

    Чому relay а не extension:
      - Chrome 130+ вимкнув Manifest V2 extensions за замовчуванням
      - Chrome не підтримує user:pass@ в --proxy-server (ERR_NO_SUPPORTED_PROXIES)
      - Relay стартує як Python thread ПЕРЕД Chrome — авторизація 100% надійна
    """
    options = ChromiumOptions()
    options.set_argument("--disable-infobars")
    options.set_argument("--no-first-run")
    options.set_argument("--no-default-browser-check")

    # Захист від WebRTC leak (без --disable-webrtc щоб не ламати сайти)
    options.set_pref("webrtc.ip_handling_policy", "disable_non_proxied_udp")
    options.set_pref("webrtc.multiple_routes_enabled", False)
    options.set_pref("webrtc.nonproxied_udp_enabled", False)

    _proxy_data = proxy_manager.load()
    use_proxy   = _proxy_data.get("use_proxy", False)
    proxies_dict = _proxy_data.get("proxies", {})

    available_proxies: List[dict] = []
    if use_proxy:
        if isinstance(proxies_dict, dict):
            available_proxies = proxies_dict.get(site_key, []) or proxies_dict.get("General", [])
        elif isinstance(proxies_dict, list):
            available_proxies = proxies_dict

    relay: Optional[LocalProxyRelay] = None

    if use_proxy and available_proxies:
        p = random.choice(available_proxies)
        host = p["host"]
        port = int(p["port"])
        user = p.get("user", "")
        password = p.get("pass", "")

        # Запускаємо relay ДО старту Chrome
        relay = LocalProxyRelay(host, port, user, password)
        local_port = relay.start()

        # Chrome підключається до localhost без пароля
        options.set_proxy(f"http://127.0.0.1:{local_port}")

        logger.info("[Proxy %s] Relay 127.0.0.1:%d → %s:%d (user: %s)",
                    site_key, local_port, host, port, user or "—")
    else:
        logger.info("Проксі для %s відсутні або вимкнені.", site_key)

    # ── Buster captcha extension ────────────────────────────────────────────
    buster_path = os.path.abspath('buster_ext')
    if os.path.exists(buster_path):
        options.add_extension(buster_path)

    try:
        return ChromiumPage(addr_or_opts=options), relay
    except Exception as e:
        logger.error("Помилка запуску браузера: %s", e)
        status_dict['is_running'] = False
        if relay: relay.stop()
        return None, None


def check_and_wait_for_captcha(page: ChromiumPage) -> None:
    wait_time = 0
    while True:
        try:
            html = str(getattr(page, 'html', '') or '').lower()

            is_captcha = (
                "oups..." in html
                or "votre réseau semble beaucoup utiliser" in html
                or (
                    ("just a moment..." in html or "cloudflare" in html)
                    and (page.eles('css:iframe[src*="cloudflare"]') or page.eles('css:iframe[src*="challenge"]'))
                )
            )

            if not is_captcha:
                break

            if wait_time == 0:
                logger.warning("CAPTCHA виявлено. Спроба автопроходження...")

            try:
                recaptcha_frame = page.get_frame('@src^https://www.google.com/recaptcha/api2/anchor')
                if recaptcha_frame:
                    checkboxes = recaptcha_frame.eles('.recaptcha-checkbox-border', timeout=1)
                    if checkboxes:
                        checkboxes[0].click()  # type: ignore[index,union-attr]
                        time.sleep(2)

                bframe = page.get_frame('@src^https://www.google.com/recaptcha/api2/bframe')
                if bframe:
                    buster_btns = bframe.eles('#solver-button', timeout=1)
                    if buster_btns:
                        buster_btns[0].click()  # type: ignore[index,union-attr]
                        time.sleep(5)

                for btn in page.eles('tag:button'):  # type: ignore[union-attr]
                    if "valider" in str(btn.text).lower() or btn.attr('type') == 'submit':
                        btn.click()
                        time.sleep(3)
                        break
            except Exception:
                pass

            if wait_time >= CAPTCHA_MAX_WAIT_SEC:
                raise Exception("NEED_PROXY_CHANGE")

            time.sleep(CAPTCHA_POLL_INTERVAL_SEC)
            wait_time += CAPTCHA_POLL_INTERVAL_SEC

        except Exception as e:
            if "NEED_PROXY_CHANGE" in str(e):
                raise
            time.sleep(2)


# ─────────────────────────────────────────────
#  ДОПОМІЖНІ ФУНКЦІЇ (усунення дублювання)
# ─────────────────────────────────────────────

def _get_link_key(item: dict) -> str:
    """Визначає правильний ключ посилання залежно від скрапера."""
    for key in ("Statement of Information (Link)", "Посилання на PDF", "Посилання"):
        if key in item:
            return key
    return ""


def _persist_result(item: dict, site_key: str, collected_data: List[dict]) -> None:
    """
    Зберігає один результат у базу + Google Sheets (якщо не дублікат).
    Єдина точка збереження замість 6 однакових блоків у run_scraping.
    """
    name = item.get("Назва")
    link_key = _get_link_key(item)
    link = item.get(link_key, "")

    if not name or database.is_company_name_scraped(name):
        logger.debug("⏩ Пропущено (дублікат або порожня назва): %s", name)
        return

    database.save_company_to_db(name, link, site_key)

    # Sheets — через чергу: послідовно, з паузою, гарантовано дійде
    _enqueue_sheet_write(name, link, site_key)

    collected_data.append(item)


def _run_simple_scraper(scraper_fn, args_builder, page, keyword: str,
                        max_count: int, site_key: str, status_dict: dict,
                        chat_id: int, collected_data: List[dict], file_format: str) -> bool:
    """
    Запускає скрапер (підтримує кілька ключових слів через кому).
    Повертає True — сигнал для run_scraping завершити роботу.

    Multi-keyword: "tech, software, startup" → три запити, результати об'єднуються.

    Ліміт для кожного ключового слова = залишок до max_count.
    Якщо перше слово дало лише 42 з 125 — наступне може заповнити решту.
    """
    keywords = [k.strip() for k in str(keyword).split(',') if k.strip()]

    for kw in keywords:
        if not status_dict.get('is_running', True):
            break

        remaining = max_count - len(collected_data)
        if remaining <= 0:
            break

        if len(keywords) > 1:
            logger.info("Multi-keyword [%s]: запит '%s' (ліміт %d, зібрано %d/%d)",
                        site_key, kw, remaining, len(collected_data), max_count)

        args    = args_builder(page, kw, remaining, status_dict)
        results = scraper_fn(*args)

        for item in results:
            if len(collected_data) >= max_count:
                break
            if not status_dict.get('is_running', True):
                break
            _persist_result(item, site_key, collected_data)

        # ── Оновлюємо загальний лічильник після кожного ключового слова ──
        # (scrapers оновлюють status_dict["current"] своїм per-keyword counter,
        #  тут скидаємо на реальний загальний підсумок)
        status_dict["current"] = len(collected_data)

    if collected_data:
        save_scraping_results(chat_id, collected_data, file_format, status_dict)
    return True


# scrape_france() видалено — France тепер використовує scrape_france_api()
# з france_scraper.py (JSON API pappers.ai, без браузера)


def scrape_finland(page: ChromiumPage, config: dict, keyword: str, seen_links: Set[str]) -> List[str]:
    url = config["search_url"].format(kw=keyword)
    logger.info("Фінляндія: %s", url)
    page.get(url)
    check_and_wait_for_captcha(page)

    unique_results: List[str] = []
    try:
        for _ in range(ELEMENT_WAIT_RETRIES):
            if page.ele('tag:table'):
                break
            time.sleep(1)
        else:
            logger.warning("Таблицю Фінляндії не знайдено.")
            return unique_results

        page.scroll.to_bottom()
        time.sleep(1.5)

        btn = (
            page.ele('css:button[aria-label="kaikki"]')
            or page.ele('css:button[aria-label="all"]')
            or page.ele('text:kaikki')
            or page.ele('text:all')
        )
        if btn:
            btn.click(by_js=True)  # type: ignore[call-arg]
            time.sleep(5)

        for el in page.eles(f'css:{config["link_selector"]}'):  # type: ignore[union-attr]
            href = getattr(el, 'attr', lambda x: None)("href") if hasattr(el, 'attr') else getattr(el, 'link', None)
            if href:
                clean = str(href).split('#')[0].split('?')[0]
                if clean not in seen_links and clean not in unique_results:
                    unique_results.append(clean)

        logger.info("Фінляндія: %d посилань", len(unique_results))
    except Exception as e:
        logger.error("Помилка Фінляндія: %s", e)

    return unique_results


def _format_excel(tmp_path: str) -> None:
    """Форматує Excel-файл: кольоровий заголовок, авто-ширина, фільтр, зебра."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        # ── Стиль заголовка (темно-синій фон, білий жирний текст) ──
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align

        # ── Зебра: кожен 2-й рядок — світло-блакитний ──
        stripe_fill = PatternFill("solid", fgColor="EBF5FB")
        for row_idx in range(2, ws.max_row + 1, 2):
            for cell in ws[row_idx]:
                if cell.fill.patternType == "none" or not cell.fill.fgColor.rgb:
                    cell.fill = stripe_fill

        # ── Авто-ширина колонок (максимум 60 символів) ──
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in column_cells:
                try:
                    val_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, val_len)
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # ── Заморожений перший рядок + авто-фільтр ──
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        wb.save(tmp_path)
    except ImportError:
        logger.warning("openpyxl не встановлений — форматування Excel пропущено.")
    except Exception as e:
        logger.warning("Помилка форматування Excel: %s", e)


def save_scraping_results(chat_id: int, data: List[dict], file_format: str, status_dict: dict) -> None:
    df = pd.DataFrame(data)
    ext = {"EXCEL": "xlsx", "JSON": "json"}.get(file_format, "txt")
    tmp_path = os.path.join(tempfile.gettempdir(), f"res_{chat_id}.{ext}")

    if file_format == "EXCEL":
        df.to_excel(tmp_path, index=False, engine="openpyxl")
        _format_excel(tmp_path)
    elif file_format == "JSON":
        df.to_json(tmp_path, orient="records", force_ascii=False, indent=4)
    else:
        df.to_csv(tmp_path, index=False, sep='\t')

    status_dict['file_path'] = tmp_path
    logger.info("Файл збережено: %s", tmp_path)


def run_scraping(chat_id: int, keyword: str, max_count: int,
                 site_key: str, file_format: str, status_dict: dict) -> None:
    database.init_db()
    config = SCRAPER_CONFIG.get(site_key, {})
    collected_data: List[dict] = []

    # ── Таблиця маршрутизації скраперів ──
    # Формат: site_key -> (scraper_fn, args_builder)
    # args_builder(page, keyword, max_count, status_dict) → tuple аргументів
    #
    # API-скрапери (France, Finland, Latvia, UK) — page=None, браузер не запускається.
    # Браузерні (California, Denmark, Czech, NZ, Thailand, Turkey) — page=ChromiumPage.
    SIMPLE_SCRAPERS: Dict[str, Any] = {
        # ── API-скрапери (без браузера) ──
        "France":        (scrape_france_api,   lambda p, kw, mc, sd: (kw, mc, sd)),
        "Finland":       (scrape_finland_api,  lambda p, kw, mc, sd: (kw, mc, sd)),
        "Latvia":        (scrape_latvia,        lambda p, kw, mc, sd: (kw, mc, sd)),
        "UnitedKingdom": (scrape_uk_api,        lambda p, kw, mc, sd: (kw, mc, sd)),
        # ── Браузерні скрапери ──
        "California":    (scrape_california,   lambda p, kw, mc, sd: (p, kw, mc, sd)),
        "Denmark":       (scrape_denmark,       lambda p, kw, mc, sd: (p, kw, mc, sd)),
        "CzechRepublic": (scrape_czech,         lambda p, kw, mc, sd: (p, kw, mc, sd)),
        "NewZealand":    (scrape_new_zealand,   lambda p, kw, mc, sd: (p, kw, mc, sd)),
        "Thailand":      (scrape_thailand,      lambda p, kw, mc, sd: (p, kw, mc, sd)),
        "Turkey":        (turkey_scraper.scrape_turkey, lambda p, kw, mc, sd: (p, kw, mc, sd)),
    }

    # Браузер запускаємо ТІЛЬКИ для скраперів, яким він потрібен
    BROWSER_BASED: set = {"California", "Denmark", "CzechRepublic", "NewZealand", "Thailand", "Turkey"}
    needs_browser = site_key in BROWSER_BASED

    page, relay = (None, None)
    if needs_browser:
        page, relay = get_page(chat_id, status_dict, site_key)
        if page is None:
            return

    try:
        if site_key in SIMPLE_SCRAPERS:
            scraper_fn, args_builder = SIMPLE_SCRAPERS[site_key]
            _run_simple_scraper(
                scraper_fn, args_builder, page, keyword, max_count,
                site_key, status_dict, chat_id, collected_data, file_format
            )
            return

        # --- Фінляндія (пагінація браузером + вкладки) ---
        # France перенесено до SIMPLE_SCRAPERS (API pappers.ai, без браузера)
        seen_names_session: Set[str] = set()
        seen_links: Set[str] = set()

        while len(collected_data) < max_count:
            if not status_dict.get('is_running', True):
                break

            try:
                valid_hrefs = scrape_finland(page, config, keyword, seen_links)  # type: ignore[arg-type]
            except Exception as e:
                if "NEED_PROXY_CHANGE" in str(e):
                    logger.info("Зміна проксі...")
                    page.quit()  # type: ignore[union-attr]
                    page, relay = get_page(chat_id, status_dict, site_key)
                    if page is None:
                        break
                    continue
                break

            if not valid_hrefs:
                break

            for link in valid_hrefs:
                if len(collected_data) >= max_count:
                    break
                if not status_dict.get('is_running', True):
                    break
                if link in seen_links:
                    continue
                seen_links.add(link)

                if database.is_company_scraped(link):
                    continue

                company_tab = None
                try:
                    company_tab = page.new_tab(link)  # type: ignore[union-attr]
                    check_and_wait_for_captcha(company_tab)

                    name_ele = None
                    for _ in range(10):
                        name_ele = company_tab.ele(f'tag:{config["name_tag"]}')
                        if name_ele:
                            break
                        time.sleep(1)

                    if not name_ele:
                        company_tab.close()
                        continue

                    raw_text = getattr(name_ele, 'text', '')
                    if callable(raw_text):
                        raw_text = raw_text()
                    name = str(raw_text).strip().split('\n')[0]

                    if not name:
                        company_tab.close()
                        continue

                    name_lower = name.lower()
                    if name_lower in seen_names_session or database.is_company_name_scraped(name):
                        company_tab.close()
                        continue
                    seen_names_session.add(name_lower)

                    item = {"Назва": name, "Посилання": link}
                    _persist_result(item, site_key, collected_data)
                    status_dict['current'] = len(collected_data)
                    status_dict['last_name'] = name
                    logger.info("%d. %s", len(collected_data), name)
                    company_tab.close()

                except Exception as e:
                    if company_tab:
                        try:
                            company_tab.close()
                        except Exception:
                            pass
                    if "NEED_PROXY_CHANGE" in str(e):
                        page.quit()  # type: ignore[union-attr]
                        page, relay = get_page(chat_id, status_dict, site_key)
                        if page is None:
                            break
                        break
                    elif "disconnected" in str(e).lower():
                        status_dict['is_running'] = False
                        break
                    else:
                        logger.error("Помилка компанії %s: %s", link, e)

            # Finland повертає всі результати за один запит — виходимо після першого pass
            break

        if collected_data:
            save_scraping_results(chat_id, collected_data, file_format, status_dict)

    finally:
        status_dict['is_running'] = False
        if page is not None:
            try:
                page.quit()  # type: ignore[union-attr]
            except Exception:
                pass
        # Зупиняємо локальний проксі-relay
        if relay is not None:
            try:
                relay.stop()
            except Exception:
                pass
        # Чекаємо поки всі записи в Google Sheets дійдуть
        logger.info("[%d] Очікую завершення запису в Sheets...", chat_id)
        flush_sheets_queue()
        logger.info("[%d] Скрапер завершив роботу.", chat_id)